"""
build_sources.py — собирает sources.json из «Радар_источники_v2.xlsx».

Запускается вручную, когда база источников обновилась:
    python build_sources.py Радар_источники_v2.xlsx

Что делает помимо конвертации:
  1. НОРМАЛИЗУЕТ метки сфер. В исходной таблице часть меток в листе «Источники»
     не совпадает с листом «Сферы» (регистр, «ё», разные формулировки) — из-за
     этого 110 из 276 источников были недостижимы для маршрутизации.
     См. SPHERE_ALIASES ниже.
  2. Отбрасывает то, что физически нельзя спарсить (нет ссылки, приватный инвайт).
  3. Приводит ссылки Telegram к виду https://t.me/<handle> и достаёт handle.

Результат — sources.json, который читает radar_routing.py в рантайме
(openpyxl и xlsx в проде не нужны).
"""
from __future__ import annotations

import json
import re
import sys
from typing import Any

from openpyxl import load_workbook

# Метки сфер из листа «Источники» → канонические сферы из листа «Сферы».
# Слева то, что реально написано в таблице; справа — то, с чем матчатся ключевые слова.
SPHERE_ALIASES: dict[str, str] = {
    "удалёнка": "удаленка",
    "remote-first компании": "удаленка",
    "номады и релокация": "релокация и удаленка",
    "фриланс или шабашки": "фриланс",
    "стартапы и it": "IT",
    "контент, маркетинг, креатив": "маркетинг",
    "юриспруденция и финансы": "юриспруденция",
    "рф и снг": "общее",
    "глобальные платформы (база)": "общее",
    "любая": "общее",
    "сайт": "общее",
}

TG_RE = re.compile(r"t\.me/(?:s/)?([A-Za-z0-9_]{4,})")


def norm_sphere(raw: str | None) -> str:
    if not raw:
        return "общее"
    key = str(raw).strip().lower()
    return SPHERE_ALIASES.get(key, str(raw).strip())


def parse_keywords(raw: str | None) -> list[str]:
    if not raw or str(raw).strip().startswith("—"):
        return []
    return [w.strip().lower() for w in str(raw).split(",") if w.strip()]


def is_parseable(row: dict[str, Any]) -> bool:
    """Физически ли можно забрать вакансии автоматом."""
    flag = str(row.get("parses") or "").strip().lower()
    if flag.startswith("нет"):
        return False
    link = str(row.get("link") or "")
    if not link.startswith("http"):
        return False          # «Название без ссылки»
    if "t.me/+" in link:
        return False          # приватный инвайт — нужна подписка аккаунта
    if "linkedin.com" in link:
        return False          # ToS + техническая блокировка
    return True


def build(xlsx_path: str, out_path: str = "sources.json") -> None:
    wb = load_workbook(xlsx_path, read_only=True)

    # --- лист «Сферы»: ключевые слова ---
    spheres: list[dict[str, Any]] = []
    for r in wb["Сферы"].iter_rows(min_row=2, values_only=True):
        name, kw = r[0], r[1]
        if not name or kw is None:
            continue
        keywords = parse_keywords(kw)
        spheres.append({"name": str(name).strip(), "keywords": keywords})

    # --- лист «Источники» ---
    telegram: list[dict[str, Any]] = []
    search: list[dict[str, Any]] = []
    manual: list[dict[str, Any]] = []

    for r in wb["Источники"].iter_rows(min_row=2, values_only=True):
        if not r[1]:
            continue
        row = {
            "type": (r[0] or "").strip(),
            "link": str(r[1]).strip(),
            "template": (r[2] or "").strip() if r[2] else "",
            "sphere": norm_sphere(r[3]),
            "level": (r[4] or "").strip() if r[4] else "",
            "region": (r[5] or "").strip() if r[5] else "",
            "priority": (r[6] or "").strip() if r[6] else "",
            "parses": r[7],
            "comment": (r[8] or "").strip() if r[8] else "",
        }

        # Поисковые источники с шаблоном подстановки.
        # Здесь «ссылка» — это название («hh.ru (API)»), поэтому пригодность
        # определяем по флагу «Парсится», а не по виду ссылки.
        if row["template"]:
            flag = str(row["parses"] or "").strip().lower()
            parseable = (not flag.startswith("нет")
                         and "linkedin" not in row["template"].lower())
            search.append({
                "name": row["link"],
                "template": row["template"],
                "sphere": row["sphere"],
                "parseable": parseable,
                "comment": row["comment"],
            })
            continue

        if not is_parseable(row):
            manual.append({
                "name": row["link"],
                "sphere": row["sphere"],
                "reason": "нет ссылки" if not row["link"].startswith("http")
                          else "приватный инвайт" if "t.me/+" in row["link"]
                          else "не парсится",
            })
            continue

        m = TG_RE.search(row["link"])
        if m:
            telegram.append({
                "handle": m.group(1),
                "url": f"https://t.me/{m.group(1)}",
                "sphere": row["sphere"],
                "level": row["level"],
                "priority": row["priority"],
            })
        else:
            # джоб-борд со ссылкой, но без шаблона поиска — отдаём человеку
            manual.append({
                "name": row["link"],
                "sphere": row["sphere"],
                "reason": "джоб-борд без шаблона поиска",
            })

    # дедуп телеграма по handle
    seen: set[str] = set()
    tg_unique = []
    for t in telegram:
        key = t["handle"].lower()
        if key in seen:
            continue
        seen.add(key)
        tg_unique.append(t)

    data = {
        "spheres": spheres,
        "telegram": tg_unique,
        "search": search,
        "manual": manual,
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    # --- отчёт ---
    declared = {s["name"] for s in spheres}
    used = {t["sphere"] for t in tg_unique} | {s["sphere"] for s in search}
    orphan = used - declared

    print(f"Сфер: {len(spheres)}")
    print(f"Telegram-каналов пригодных: {len(tg_unique)}")
    print(f"Поисковых источников: {len(search)}")
    print(f"Отдаём ссылками человеку: {len(manual)}")
    if orphan:
        print(f"ВНИМАНИЕ, сферы без ключевых слов: {sorted(orphan)}")
    else:
        print("Все сферы источников сопоставлены с ключевыми словами ✔")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "Радар_источники_v2.xlsx"
    build(path)
